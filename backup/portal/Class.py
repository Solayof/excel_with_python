#!/usr/bin/python3
"""class model
"""
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from sqlalchemy import Column, ForeignKey, String
from sqlalchemy.orm import relationship
from backup.base import Base
from backup.engine.dbSheet import Workbook
from backup.baseModel import BaseModel

class Class(BaseModel, Base):
    """class model
    
    Usage: jss1 = Class(className="Jss 1")
            others parameters are optional

    Args:
        BaseModel (_type_): Basemodel class
        Base (_type_): declarative base
    """    
    __tablename__ = "classes"
    extend_existing = True
    code = Column(String(16), unique=True)
    arm = Column(String(1))
    session = Column(String(9))
    className = Column(String(6)) 
    students = relationship(
        "Student",
        foreign_keys="[Student.classroom_id]",
        back_populates="classroom",
        uselist=True
        )

    def __init__(self, *args, **kwargs):
        """initializing class
        """
        className = None       
        if kwargs:
            className = kwargs.pop("className", None)
            arm = kwargs.pop("arm", None)

        if className:
            className = className.upper() + arm.upper()
            code = className.replace(" ", "-")

            yr = int(datetime.now().strftime("%y"))
            mth = datetime.now().strftime("%m")
            if int(mth) < 8:
                yr = yr - 1
            kwargs["code"] = code + "-" + f"20{yr}-20{yr + 1}"
            kwargs["className"] = className
            kwargs["arm"] = arm
            super().__init__(*args, **kwargs)

    def save(self):
        """class save method
        """        
        if not self.session:
            yr = datetime.now().strftime("%y")
            self.session = f"20{yr}-20{int(yr) + 1}"
        super().save()
        
    def to_dict(self):
        """dictionary representation of class instance

        Returns:
            _type_: dict
        """        
        new_dict = self.__dict__.copy()
        new_dict.pop("_sa_instance_state", None)
       
        new_dict["created_at"] = self.created_at.isoformat()
        new_dict["updated_at"] = self.updated_at.isoformat()
        
        students = self.students
        new_dict["number_of_students"] = len(students) if students else 0
        
        
        
        return new_dict
    
    def getSheet(self):
        worksheet = Workbook(filePath=f"{self.code}.xlsx", defaultFile=f"{self.className[:3]}.xlsx")
        worksheet.open_session()

        return worksheet
    
    def generateSheet(self):
        students = self.students
        students.sort(key=lambda s: s.fullName)
        worksheet = self.getSheet()
        worksheet.open_session()
        sheet = worksheet.getDbsheet()

        stdcell = "b4"
        stdrow, stdcol = coordinate_to_tuple(stdcell)
        for student in students:
            sheet.cell(stdrow, stdcol, student.fullName)
            subjects = student.subjects
    
            sheet.cell(stdrow, stdcol + 1, student.admission_no)

            
            sheet.cell(stdrow, stdcol + 2, student.gender)
            
            sheet.cell(stdrow, stdcol + 3, student.classroom.className)
            for subject in subjects:
                subCell = worksheet.getSubjectCell(subject=subject.name)
                _, subcol = coordinate_to_tuple(subCell)

                sheet.cell(stdrow, subcol, subject.CA)

                sheet.cell(stdrow, subcol + 1, subject.examScore)

                sheet.cell(stdrow, subcol + 3, subject.secondTermScore)

                sheet.cell(stdrow, subcol + 4, subject.firstTermScore)
                
            stdrow = stdrow + 1
        worksheet.saveWorkbook(self.code)

    @property
    def sheetSubjects(self):
        return self.getSheet().dbSubjects()
    

    def students_to_dict(self):
        """dictionary representation of class instance

        Returns:
            _type_: dict
        """
        students = self.students
        students.sort(key=lambda s: s.fullName)
        new_dict = [{"FULLNAME": std.fullName, "NUMBER OF SUBJECTS RECORDED": len(std.subjects)} for std in students]

        return new_dict
    
    # def getStudentsFullName(self):
    #     students = self.students
    #     students.sort(key=lambda s: s.fullName)
    #     return [std.fullName for std in students]
    
    def getStudentsIdandNames(self):
        students = self.students
        students.sort(key=lambda s: s.fullName)
        return {std.fullName: std.id for std in students}
        

        


