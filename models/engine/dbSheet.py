from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter


class Workbook:
    __workbook = None
    filePath = ''
    
 
    def __init__(self, *args, **kwargs):
        self.filePath = kwargs.pop("filePath", None)
        self.defaultFile =  kwargs.pop("defaultFile", None)
    
    def open_session(self):
        if not self.__workbook:
            try:
                self.__workbook = load_workbook(self.filePath)
            except FileNotFoundError:
                self.__workbook = load_workbook(self.defaultFile, read_only=False)

       
    def getDbsheet(self, sheetName):
        return self.__workbook[sheetName]

    def saveWorkbook(self, fileName):
        if self.__workbook is None:
            return False
        self.__workbook.save(f"{fileName}.xlsx")
        return True

    def dbSubjects(self, sheetName="1ST TERM Db"):
        sub = []
        cell = "K2"
        db = self.getDbsheet(sheetName)
        while db[cell].value:
            sub.append(db[cell].value)
            row, col = coordinate_to_tuple(cell)
            cell = f"{get_column_letter(col + 3)}{row}"
            print(f"cell: {cell}")
        return sub

    def free_subject_cells(self, sheetName):
        subs = {}
        cell = 'K2'
        db = self.getDbsheet(sheetName)
        while db[cell].value:
            if db[cell].value.startswith("="):
                pass
    
    def writeCell(self, cell, value, sheetName="1ST TERM Db"):
        sheet = self.getDbsheet(sheetName=sheetName)
        sheet[cell] = value

    def readCell(self, cell, sheetName="1ST TERM Db"):
        sheet = self.getDbsheet(sheetName=sheetName)
        return sheet[cell].value
    
    def getSubjectCell(self, subject, sheetName="1ST TERM Db"):
        cell = "K2"
        db = self.getDbsheet(sheetName)
        while db[cell].value and db[cell].value != subject:
            row, col = coordinate_to_tuple(cell)
            cell = f"{get_column_letter(col + 3)}{row}"

        if db[cell].value == subject:
            return cell
    def get_subject_dict(self, sheetName):
        sub_cells = {}
        scell = 'CK4'
        db = self.getDbsheet(sheetName=sheetName)
        cell = "K2"
        while db[cell].value:
            db_cell = db[cell]
            if db_cell.value.startswith('='):
                sub_cells[f'{db[db_cell.value[1:]].value}_{db_cell.coordinate}'] = db_cell.coordinate
            else:
                sub_cells[db_cell.value] = db_cell.coordinate
            row, col = coordinate_to_tuple(cell)
            cell = f"{get_column_letter(col + 3)}{row}"

        return sub_cells


    def close(self):
        self.__workbook.close()


    def reload(self):
        self.__workbook = load_workbook(self.filePath)